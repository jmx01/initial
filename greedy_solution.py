import time
import numpy as np
import pandas as pd
import copy
import openpyxl
import random


def primary_deal_npz(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    max_columns = sheet.max_column
    column = openpyxl.utils.get_column_letter(max_columns)

    all_list = []
    each_list = []

    i = 1
    str1 = 'A' + str(i)
    str2 = '%s' + str(i)

    while sheet[str1:str2 % column][0][0].value is not None:
        each_list.append(sheet[str1:str2 % column][0][0].value)
        each_list.append([])  # 一个空列表后续存禁接区长度和性质

        for j in range(len(sheet[str1:str2 % column][0]) - 1):
            length = sheet[str1:str2 % column][0][j + 1].value
            str1 = 'A' + str(i + 1)
            str2 = '%s' + str(i + 1)
            pro = sheet[str1:str2 % column][0][j + 1].value
            str1 = 'A' + str(i)
            str2 = '%s' + str(i)
            each_list[1].append([length, pro])

        all_list.append(each_list)

        i += 2
        str1 = 'A' + str(i)
        str2 = '%s' + str(i)
        each_list = []

    return all_list


def standard_data_input(table):
    out_table = table.copy(deep=True)  # 深复制镜像
    out_table = out_table.groupby("长度").sum()  # 合并长度相同的行，index变为长度
    out_table["长度"] = out_table.index
    out_table.index = range(len(out_table["长度"]))  # 重新变为[编号，数量，长度]的形式
    return out_table


def alpha_effect(table, al):
    for i in range(len(table)):
        for j in range(len(table[i][1])):
            if j not in [0, len(table[i][1]) - 1]:
                if table[i][1][j][1] == 1:
                    table[i][1][j][0] += 2 * al
                else:
                    table[i][1][j][0] -= 2 * al
            else:
                if table[i][1][j][1] == 1:
                    table[i][1][j][0] += al
                else:
                    table[i][1][j][0] -= al

    for i in range(len(table)):
        j = 0
        while j < len(table[i][1]) - 1:
            if table[i][1][j][0] < 0:
                if j != 0:
                    table[i][1][j][0] = table[i][1][j][0] + table[i][1][j + 1][0] + table[i][1][j - 1][0]
                    del table[i][1][j + 1], table[i][1][j - 1]
                    j -= 1
                else:
                    table[i][1][j][0] = table[i][1][j][0] + table[i][1][j + 1][0]
                    del table[i][1][j + 1]
            else:
                j += 1

    return table


def standard_no_pick_zone(table, al):
    """
    :aid 处理禁接区（标准化禁接区）
    :param al:
    :param table: 输入的原始的禁接矩阵
    :return: 处理好后的标准禁接矩阵
    """
    for i in range(len(table)):
        j = 0
        while j < len(table[i][1]) - 1:
            if table[i][1][j][1] == table[i][1][j + 1][1]:
                table[i][1][j][0] += table[i][1][j + 1][0]
                del table[i][1][j + 1]
            else:
                j += 1

    table = alpha_effect(table, al)
    zone = np.array(table)
    calculate_zone = copy.deepcopy(table)
    for i in range(len(table)):
        for j in range(len(table[i][1]) - 1):
            calculate_zone[i][1][j + 1][0] += calculate_zone[i][1][j][0]
    return zone, calculate_zone


def e_rule(string, e):
    """
    :param e: 概率
    :param string: 一个字符
    :return: 按规则返回原字符或相反的字符
    """
    number = random.random()
    if string == "max":
        if number < e:
            return "min"
        else:
            return string
    else:
        if number < e:
            return "max"
        else:
            return string


def MM_fetchOne(table, MaxOrMin="max", e=0.05):
    """
    默认table不为空
    最大或最小的一根  # 默认取最大的一根
    :param e: 概率
    :param MaxOrMin: 标识符
    :param table: 原料或零件的镜像
    :return: 新的table和那根材料的名称数量和长度
    """
    fetch = [0, 0, 0]  # 初始化取的材料
    address = 0  # 初始化取的材料的位置
    table = np.array(table)
    quantity = table[:, 1].copy()
    length = table[:, 2].copy()
    MaxOrMin = e_rule(MaxOrMin, e)

    if MaxOrMin == "max":  # 取最大的一根
        for i in range(len(quantity)):
            if quantity[i] > 0 and length[i] > fetch[2]:
                address = i
                fetch[2] = length[i]
    else:  # 取最小的一根
        fetch[2] = 1000000  # 一个很大的常数,用来方便后续处理
        for i in range(len(quantity)):
            if quantity[i] > 0 and length[i] < fetch[2]:
                address = i
                fetch[2] = length[i]

    fetch = table[address, :].copy()  # fetch为复制的table第i+1行的信息
    fetch[1] = 1  # 拿走的数量为1
    table[address, 1] -= 1  # 数量-1
    return table, fetch


def is_continue(dt):
    """
    判断零件管是否用完
    :param dt: 零件管的镜像，一直在改变的值
    :return: 0或1
    """
    dt = pd.DataFrame(dt)
    for i in dt.loc[:, 1]:
        if i > 0:
            return 1  # 未用完
    else:
        return 0  # 已用完


def change_solution(GS_new, string):
    if string == "meng":
        GM = copy.deepcopy(GS_new)
        meng_1 = [[0], []]
        meng_2 = [0]
        meng_3 = []
        ma_in = pd.DataFrame(GM[0]).iloc[:, [0, 2]]

        # 计算 坐标，长度
        ma = np.array(ma_in).tolist()
        ma.insert(0, [0, [0]])
        for i in range(ma_in.shape[0]):
            ma[i + 1][0] = ma[i + 1][0] + ma[i][0]

        for i in range(ma_in.shape[0]):
            for j in ma[i + 1][1]:
                meng_1[0].append(j + ma[i][0])
        meng_1[0].pop()
        for i in range(len(meng_1[0])):
            meng_1[1].append(GM[1][i][1])

        # 计算 触发禁忌的管材
        ma.pop(0)
        ma_in = pd.DataFrame(GM[0])
        ma = np.array(ma_in).tolist()
        for i in ma:
            if i[4] > 0:
                i.append(1)
            else:
                i.append(0)

            i.append(len(i[2]))

        for i in range(ma_in.shape[0] - 1):
            ma[i + 1][6] = ma[i + 1][6] + ma[i][6]
            if ma[i + 1][5] == 1:
                meng_2.append(ma[i + 1][6] + 1)

        GS = [meng_1, meng_2, meng_3]
        return GS
    elif string == "yuan":
        GY = copy.deepcopy(GS_new)
        ma_in = copy.deepcopy(GY[0])
        GC = []

        for i in ma_in:
            i.append(len(i[2]))
        for i in range(len(ma_in) - 1):
            ma_in[i + 1][5] = ma_in[i + 1][5] + ma_in[i][5]
        ma_in.insert(0, [0, 0, [0], 0, 0, 0])
        po = GY[1]
        for i in range(len(ma_in) - 1):
            each = []
            last_count = ma_in[i][5]  # 上一根原料到的根数。
            for j in range(len(ma_in[i + 1][2])):
                if ma_in[i + 1][2][j] == po[last_count + j][1]:
                    each.append([1, last_count + j, po[last_count + j][1]])
                else:
                    each.append([2, last_count + j, ma_in[i + 1][2][j]])
            if ma_in[i + 1][3] != 0:
                each.append([2, ma_in[i + 1][5], ma_in[i + 1][3]])
            each.append([0, ma_in[i + 1][4]])
            GC.append(each)

        return GC


class greedy_solve(object):
    greedy_solution_quantity = 2  # 需要的贪婪解初始数
    over_time = 3600  # 初始解生成时间限制
    pick_up = 30  # 可放弃的新生成的原材料长度
    alpha = 0  # 建议离禁焊区的距离
    e = 0.05  # 概率随机取值
    datatable_input = 'data_input.xlsx'  # 输入文件的路径
    datatable_output = 'data_output.xlsx'  # 零件文件路径
    deal_no_pick_zone = 'zone.xlsx'  # 禁接区文件路径

    datatable_input = pd.read_excel(datatable_input)  # [编号、数量、长度]  输入材料
    datatable_output = pd.read_excel(datatable_output)  # [编号、数量、长度]  输出材料
    material_length = sum(np.array(datatable_input.iloc[:, 1]) * np.array(datatable_input.iloc[:, 2]))  # 输入材料总长度
    product_length = sum(np.array(datatable_output.iloc[:, 1]) * np.array(datatable_output.iloc[:, 2]))  # 输出材料总长度

    deal_no_pick_zone = primary_deal_npz(deal_no_pick_zone)  # 将文件处理为可处理的形式
    dt_input = standard_data_input(datatable_input)  #
    no_pick_zone, calculate_no_pick_zone = standard_no_pick_zone(deal_no_pick_zone, alpha)

    def __int__(self, greedy_solution_quantity, over_time, pick_up, alpha, e, material_length, product_length, dt_input,
                no_pick_zone, calculate_no_pick_zone):

        self.greedy_solution_quantity = greedy_solution_quantity
        self.over_time = over_time
        self.pick_up = pick_up
        self.alpha = alpha
        self.e = e
        self.material_length = material_length
        self.product_length = product_length
        self.dt_input = dt_input
        self.no_pick_zone = no_pick_zone
        self.calculate_no_pick_zone = calculate_no_pick_zone

    def __generate_tube_zone(self, x):
        npz = np.array(self.no_pick_zone)

        name = npz[:, 0]
        name = int(np.where(name == x)[0])
        np_out = self.calculate_no_pick_zone[name][1]
        np_out = np.array(np_out)
        return np_out

    def __deal_in_out(self, fetch_in_res, fetch_out, k, dt_in):
        res = 0
        use_res = k  # 被使用的剩余长度
        done = 1  # 没有切割
        np_out = self.__generate_tube_zone(fetch_out[0])

        for i in range(len(np_out)):
            if np_out[i, 0] > fetch_in_res:
                if np_out[i, 1] == 0:  # 在非禁接区
                    break
                else:  # 在禁接区
                    if i > 0:
                        res = fetch_in_res - np_out[i - 1, 0]
                    else:
                        res = fetch_in_res
                    use_res = use_res - res
                    done = 0
                    break

        fetch_out[2] = fetch_out[2] + res - fetch_in_res

        if res > self.pick_up:  # 将新的解加入dt_in中
            dt_in = pd.DataFrame(dt_in)
            # add element
            dt_in.loc[dt_in.shape[0]] = [dt_in.shape[0] + 1, 1, res]

            dt_in = np.array(dt_in)

        return dt_in, use_res, fetch_out, done

    def solve(self):
        dt_in = copy.deepcopy(self.dt_input)
        dt_out = copy.deepcopy(self.datatable_output)
        ma_input = []  # [[原料长度，剩余长度,[切割长度]，剩余长度的使用长度]]
        pro_output = []  # [[零件管编号，长度]]
        time1 = time.time()  # 开启一次贪婪解的初始时间
        fetch_in_res = 0  # 原料剩下的长度
        cut_list = []  # 每根原料的切割列表

        dt_out, fetch_out = MM_fetchOne(dt_out, "min")  # 先取一个零件管，更新数量
        pro_output.append([fetch_out[0], fetch_out[2]])  # 添加到取出的零件管集中
        dt_in, fetch_in = MM_fetchOne(dt_in)  # 再取一个原料管
        ma_input.append([fetch_in[2], fetch_in[2], cut_list, 0])  # 添加到原料管集中
        fetch_in_res += fetch_in[2]

        while is_continue(dt_out) or fetch_out[2] != 0:  # 当原料管还有剩余，或者拿出的零件管长度不为0
            time2 = time.time()
            time_break = time2 - time1
            if time_break > self.over_time:
                return 0  # 单个初始解超时判别

            while fetch_in_res >= fetch_out[2]:  # 当剩余长度大于取出长度
                if len(cut_list) == 0:
                    cut_list.append(fetch_out[2] - ma_input[-1][3])  # 如果是初次切割，减去上根原料剩余长度使用部分。
                else:
                    cut_list.append(fetch_out[2])
                fetch_in_res -= fetch_out[2]
                ma_input[-1][1] -= fetch_out[2]
                fetch_out[2] = 0
                if is_continue(dt_out):  # 判断零件管是否用完
                    dt_out, fetch_out = MM_fetchOne(dt_out, "min")
                    pro_output.append([fetch_out[0], fetch_out[2]])
                else:
                    break

            while fetch_in_res < fetch_out[2]:
                ma_input[-1][2] = copy.deepcopy(cut_list)  # 把切割长度存下
                cut_list = []

                dt_in, useful_part, fetch_out, done = self.__deal_in_out(fetch_in_res, fetch_out, ma_input[-1][1],
                                                                         dt_in)
                ma_input[-1][3] = useful_part

                if is_continue(dt_in) and (sum(dt_in[:, 1]) >= 1 or done):  # 在最后一个时会循环
                    dt_in, fetch_in = MM_fetchOne(dt_in)
                    fetch_in_res = fetch_in[2]
                    ma_input.append([fetch_in[2], fetch_in[2], cut_list, 0])
                else:
                    return 1  # 原料管被用光

        for ele in ma_input:
            ele.append(ele[1] - ele[3])

        # 对原料管的处理
        dt_in = pd.DataFrame(dt_in)
        dt_in.columns = ["编号", "数量", "长度"]
        dt_in = dt_in.groupby("长度").sum()
        dt_in = dt_in[dt_in["数量"] > 0]
        dt_in["长度"] = dt_in.index.tolist()
        dt_in.index = range(dt_in.shape[0])
        dt_in.to_excel("./此次切割剩余原料.xlsx")

        # 下两行用来检测ma_input
        ma = copy.deepcopy(ma_input)
        ma = pd.DataFrame(ma, columns=["原料长度", "剩余长度", "切割向量", "剩余长度中有效使用部分", "每根原料管剩余长度"])
        ma.to_excel("./某一次的原料切割方式.xlsx")
        ma = ma.drop(ma[ma["原料长度"] == ma["每根原料管剩余长度"]].index)
        ma.to_excel("./某一次的原料切割方式——new.xlsx")

        ma_input = np.array(ma).tolist()

        return [ma_input, pro_output]
