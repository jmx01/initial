import copy
from decimal import Decimal

import numpy as np
import openpyxl
import pandas as pd

file = ["5407055-B1", "7409753", "7409762", "G11-79 7406104-E1"]


def path(f):
    di = "./file/" + f + '/data_input.xlsx'
    do = "./file/" + f + '/data_output.xlsx'
    zo = "./file/" + f + '/zone.xlsx'
    return di, do, zo


def primary_deal_npz(f):
    """读文件，生成初始禁接区"""
    wb = openpyxl.load_workbook(f)
    sheet = wb.active
    max_columns = sheet.max_column
    column = openpyxl.utils.get_column_letter(max_columns)

    all_list, each_list = [], []

    i = 1
    str1, str2 = 'A' + str(i), '%s' + str(i)

    while sheet[str1:str2 % column][0][0].value is not None:
        each_list.append(sheet[str1:str2 % column][0][0].value)
        each_list.append([])  # 一个空列表后续存禁接区长度和性质

        for j in range(len(sheet[str1:str2 % column][0]) - 1):
            length = sheet[str1:str2 % column][0][j + 1].value
            str1, str2 = 'A' + str(i + 1), '%s' + str(i + 1)
            pro = sheet[str1:str2 % column][0][j + 1].value
            str1, str2 = 'A' + str(i), '%s' + str(i)
            each_list[1].append([length, pro])

        all_list.append(each_list)

        i += 3  # 依据空格距离可修改
        str1, str2, each_list = 'A' + str(i), '%s' + str(i), []

    for i in range(len(all_list)):  # 去除尾部空
        while all_list[i][1][-1][0] is None:
            all_list[i][1].pop()

    for i in range(len(all_list)):
        for j in range(len(all_list[i][1])):
            all_list[i][1][j][0] = Decimal(str(all_list[i][1][j][0])).quantize(Decimal("0.01"),
                                                                               rounding="ROUND_HALF_UP")
    return all_list


def yuan_change(all_list, pick_up):
    new_list = copy.deepcopy(all_list)
    for i in range(len(new_list)):
        length = 0
        for j in range(len(new_list[i][1])):
            length += new_list[i][1][j][0]
        if length <= pick_up:
            new_list[i][1] = [length, 1]
        else:
            res = length - pick_up
            for j in range(len(new_list[i][1])):
                if new_list[i][1][j][0] < res:
                    res -= new_list[i][1][j][0]
                elif new_list[i][1][j][0] == res:
                    del new_list[i][1][j + 1:]
                    new_list[i][1].insert(j + 1, [pick_up, 1])
                    break
                elif new_list[i][1][j][0] > res:
                    new_list[i][1].insert(j, [res, new_list[i][1][j][1]])
                    del new_list[i][1][j + 1:]
                    new_list[i][1].insert(j + 1, [pick_up, 1])
                    break
    return new_list


def alpha_effect(table, al):
    """根据建议离禁焊区距离，扩大禁接区"""
    for i in range(len(table)):
        for j in range(len(table[i][1])):
            window = table[i][1][j]
            if j not in [0, len(table[i][1]) - 1]:
                if window[1] == 1:
                    window[0] += 2 * al
                else:
                    window[0] -= 2 * al
            else:
                if window[1] == 1:
                    window[0] += al
                else:
                    window[0] -= al

    for i in range(len(table)):
        j = 0
        while j < len(table[i][1]) - 1:
            if table[i][1][j][0] < 0:
                if j != 0:
                    table[i][1][j][0] = table[i][1][j][0] + table[i][1][j + 1][0] + table[i][1][j - 1][0]
                    table[i][1][j][1] = 1
                    del table[i][1][j + 1], table[i][1][j - 1]
                    j -= 1
                else:
                    table[i][1][j][0] = table[i][1][j][0] + table[i][1][j + 1][0]
                    table[i][1][j][1] = 1
                    del table[i][1][j + 1]
            else:
                j += 1

    return table


def standard_no_pick_zone(table, al=0):
    """
    :aid 处理禁接区（标准化禁接区）
    :param al:建议离焊点距离
    :param table: 输入的原始的禁接矩阵
    :return: 处理好后的标准禁接矩阵
    """
    table = copy.deepcopy(table)
    for i in range(len(table)):
        j = 0
        while j < len(table[i][1]) - 1:
            if table[i][1][j][1] == table[i][1][j + 1][1]:
                table[i][1][j][0] += table[i][1][j + 1][0]
                del table[i][1][j + 1]
            else:
                j += 1

    table = alpha_effect(table, al)
    zone = np.array(table, dtype=object)
    calculate_zone = copy.deepcopy(table)

    for i in range(len(table)):
        for j in range(len(table[i][1]) - 1):
            calculate_zone[i][1][j + 1][0] += calculate_zone[i][1][j][0]
    return zone, calculate_zone


def standard_data_output(table):
    for i in range(len(table)):
        table.iloc[i, 2] = Decimal(str(table.iloc[i, 2])).quantize(Decimal("0.01"), rounding="ROUND_HALF_UP")
    return table


def standard_data_input(table):
    """标准化原料管（将长度相同的管材合在一起 ）"""
    table = table.groupby("长度").sum()  # 合并长度相同的行，index变为长度
    table["长度"] = table.index
    table.index = range(len(table["长度"]))  # 重新变为[编号，数量，长度]的形式
    table = standard_data_output(table)
    return table


def seam_num(table):
    """给出产品管焊缝限制"""
    table["焊缝上限"] = " "
    for i in range(len(table)):
        if table.iloc[i][2] <= 2000:
            table.iloc[i, 3] = 0
        elif 2000 < table.iloc[i][2] <= 5000:
            table.iloc[i, 3] = 1
        elif 5000 < table.iloc[i][2] <= 10000:
            table.iloc[i, 3] = 2
        elif 10000 < table.iloc[i][2] <= 15000:
            table.iloc[i, 3] = 3
        elif 15000 < table.iloc[i][2] <= 25000:
            table.iloc[i, 3] = 4
        else:
            table.iloc[i, 3] = 4 + (table.iloc[i][2] - 25000) // 4000
    return table


class initial_data(object):
    greedy_solution_quantity = 1  # 需要的贪婪解初始数
    algorithm_solution_quantity = 1  # 组批解初始数
    over_time = 30  # 初始解生成时间限制
    pick_up = 500
    pick_up = Decimal(pick_up).quantize(Decimal("0.01"), rounding="ROUND_HALF_UP")  # 可放弃的最大材料长度
    alpha = 50  # 建议离禁焊区的距离
    e = 0.05  # 概率随机取值
    num = 1000  # 迭代次数
    connection = False  # 是否虚焊
    flag_yuan = False  # 是否使用袁浩的禁接区
    show_composition = True
    datatable_input, datatable_output, deal_no_pick_zone = path(file[2])
    # datatable_input = 'data_input.xlsx'  # 输入文件的路径
    # datatable_output = 'data_output.xlsx'  # 零件文件路径
    # deal_no_pick_zone = 'zone.xlsx'  # 禁接区文件路径

    datatable_input = pd.read_excel(datatable_input)  # [编号、数量、长度]  输入材料
    datatable_output = pd.read_excel(datatable_output)  # [编号、数量、长度、焊缝上限]  输出材料
    datatable_output = seam_num(datatable_output)
    material_length = sum(np.array(datatable_input.iloc[:, 1]) * np.array(datatable_input.iloc[:, 2]))  # 输入材料总长度
    product_length = sum(np.array(datatable_output.iloc[:, 1]) * np.array(datatable_output.iloc[:, 2]))  # 输出材料总长度

    deal_no_pick_zone = primary_deal_npz(deal_no_pick_zone)  # 将文件处理为可处理的形式
    if flag_yuan:
        new_pick_zone = yuan_change(deal_no_pick_zone, pick_up)  # 给袁浩用的新禁接区
    else:
        new_pick_zone = copy.deepcopy(deal_no_pick_zone)  # 老禁接区
    dt_input = standard_data_input(datatable_input)  #
    no_pick_zone, calculate_no_pick_zone = standard_no_pick_zone(deal_no_pick_zone, alpha)
    new_no_pick_zone, new_calculate_no_pick_zone = standard_no_pick_zone(new_pick_zone, alpha)
    datatable_output = standard_data_output(datatable_output)

    def change_zone(self, index):
        """改变函数，为类的内函数，调用即可，index是顺序"""
        change_zone = ["change", []]
        for i in range(len(index)):
            change_zone[1].extend(self.new_pick_zone[index[i]][1])
        change_zone = [change_zone]
        change_zone, change_zone_cal = standard_no_pick_zone(change_zone, self.alpha)
        new_change_zone = []
        new_change_zone_cal = change_zone_cal[0][1]
        if new_change_zone_cal[0][1] == 1:
            i = 0
        else:
            i = 1

        while i < len(new_change_zone_cal):
            if i == 0:
                new_change_zone.append([0, new_change_zone_cal[i][0]])
            else:
                new_change_zone.append([new_change_zone_cal[i - 1][0], new_change_zone_cal[i][0]])
            i = i + 2

        return new_change_zone, change_zone, change_zone_cal
